# راوتر المشرف الإداري - Morix Platform
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from app.auth import get_current_user, hash_password
from app.database import get_db
import logging
import io

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin", tags=["المشرف الإداري"])


def _require_admin(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ("admin", "manager", "owner"):
        raise HTTPException(status_code=403, detail="هذه الخدمة للمشرفين الإداريين فقط")
    return current_user


# ============================================
# نظرة عامة
# ============================================
@router.get("/overview")
async def get_overview(current_user: dict = Depends(_require_admin), db=Depends(get_db)):
    school_id = current_user.get("school_id")
    if not school_id:
        return {"total_students": 0, "total_teachers": 0, "total_homework": 0, "total_tests": 0}

    students = db.table("users").select("id").eq("role", "student").eq("school_id", school_id).execute()
    teachers = db.table("users").select("id").eq("role", "teacher").eq("school_id", school_id).execute()
    homework = db.table("homework").select("id").eq("school_id", school_id).execute()
    tests = db.table("tests").select("id").eq("school_id", school_id).execute()

    return {
        "total_students": len(students.data or []),
        "total_teachers": len(teachers.data or []),
        "total_homework": len(homework.data or []),
        "total_tests": len(tests.data or []),
    }


# ============================================
# إدارة الطلاب
# ============================================
@router.get("/students")
async def get_students(current_user: dict = Depends(_require_admin), db=Depends(get_db)):
    school_id = current_user.get("school_id")
    query = db.table("users").select(
        "id, full_name, email, grade, ministry_id, is_active, stars_count, streak_count, must_change_password, created_at"
    ).eq("role", "student")
    if school_id:
        query = query.eq("school_id", school_id)
    result = query.order("full_name").execute()
    return result.data


@router.put("/students/{student_id}/reset-password")
async def reset_password(
    student_id: str,
    body: dict,
    current_user: dict = Depends(_require_admin),
    db=Depends(get_db)
):
    new_password = body.get("new_password", "")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="كلمة المرور لازم تكون 6 أحرف على الأقل")
    if len(new_password) > 128:
        raise HTTPException(status_code=400, detail="كلمة المرور طويلة جداً (الحد الأقصى 128 حرف)")

    # التحقق أن الطالب ينتمي لنفس المدرسة (منع IDOR)
    school_id = current_user.get("school_id")
    if school_id:
        student = db.table("users").select("school_id, role").eq("id", student_id).execute()
        if not student.data:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if student.data[0].get("school_id") != school_id:
            raise HTTPException(status_code=403, detail="لا يمكنك تعديل حسابات خارج مدرستك")
        if student.data[0].get("role") in ("manager", "owner"):
            raise HTTPException(status_code=403, detail="لا يمكن إعادة تعيين كلمة سر هذا الحساب")

    new_hash = hash_password(new_password)
    db.table("users").update({
        "password_hash": new_hash,
        "must_change_password": False
    }).eq("id", student_id).execute()

    # حفظ الباسوورد الجديد في school_setup
    school_id = current_user.get("school_id")
    if school_id:
        try:
            student = db.table("users").select("full_name, email").eq("id", student_id).execute()
            if student.data:
                setup = db.table("school_setup").select("passwords_data").eq("school_id", school_id).execute()
                passwords = setup.data[0]["passwords_data"] if setup.data and setup.data[0].get("passwords_data") else []
                # تحديث كلمة مرور الطالب في القائمة
                updated = False
                for acc in passwords:
                    if acc.get("email") == student.data[0]["email"]:
                        acc["password"] = new_password
                        updated = True
                        break
                if not updated:
                    passwords.append({
                        "full_name": student.data[0]["full_name"],
                        "email": student.data[0]["email"],
                        "password": new_password,
                        "role": "student"
                    })
                db.table("school_setup").upsert({
                    "school_id": school_id,
                    "passwords_data": passwords,
                    "total_accounts_generated": len(passwords)
                }).execute()
        except Exception as e:
            logger.warning(f"Failed to save password: {e}")

    return {"message": "تم إعادة تعيين كلمة المرور بنجاح"}


@router.put("/students/{student_id}/toggle")
async def toggle_student(
    student_id: str,
    current_user: dict = Depends(_require_admin),
    db=Depends(get_db)
):
    school_id = current_user.get("school_id")
    query = db.table("users").select("is_active, school_id, role").eq("id", student_id)
    student = query.execute()
    if not student.data:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    # منع IDOR: الأدمن لا يستطيع تعطيل حسابات خارج مدرسته
    if school_id and student.data[0].get("school_id") != school_id:
        raise HTTPException(status_code=403, detail="لا يمكنك تعديل حسابات خارج مدرستك")
    if student.data[0].get("role") in ("manager", "owner"):
        raise HTTPException(status_code=403, detail="لا يمكن تعطيل حساب مدير أو مالك")
    new_status = not student.data[0]["is_active"]
    db.table("users").update({"is_active": new_status}).eq("id", student_id).execute()
    return {"message": "تم التحديث", "is_active": new_status}


# ============================================
# رفع ملف Excel لإنشاء الحسابات
# ============================================
@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    school_id: str = Form(None),
    current_user: dict = Depends(_require_admin),
    db=Depends(get_db)
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="يجب رفع ملف Excel (.xlsx أو .xls)")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة openpyxl غير مثبتة")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="تعذر قراءة الملف — تأكد أنه ملف Excel صحيح")

    students = []
    teachers = []

    # ورقة الطلاب
    student_sheet = None
    for name in ["Students", "طلاب", "students", "Sheet1"]:
        if name in wb.sheetnames:
            student_sheet = wb[name]
            break
    if not student_sheet:
        student_sheet = wb.active  # أول ورقة

    if student_sheet:
        for row in student_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                students.append({
                    "name": str(row[0]).strip(),
                    "ministry_id": str(row[1] or "000").strip(),
                    "grade": str(row[2] or "").strip(),
                })

    # ورقة المعلمين
    for name in ["Teachers", "معلمون", "معلمين", "teachers"]:
        if name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    teachers.append({
                        "name": str(row[0]).strip(),
                        "subject": str(row[1] or "").strip(),
                    })
            break

    if not students and not teachers:
        raise HTTPException(
            status_code=400,
            detail="الملف فارغ أو لا يحتوي على بيانات. الأعمدة المطلوبة: الاسم | الرقم الوزاري | الصف"
        )

    target_school = school_id or current_user.get("school_id")
    if not target_school:
        raise HTTPException(status_code=400, detail="لم يتم تحديد المدرسة")

    from app.services.account_generator import generate_accounts
    accounts = await generate_accounts(
        school_id=target_school,
        students=students,
        teachers=teachers,
        admins_count=0,
        db=db
    )

    # حفظ الباسووردات
    try:
        existing = db.table("school_setup").select("passwords_data").eq("school_id", target_school).execute()
        existing_passwords = existing.data[0]["passwords_data"] if existing.data and existing.data[0].get("passwords_data") else []
        all_passwords = existing_passwords + accounts
        db.table("school_setup").upsert({
            "school_id": target_school,
            "passwords_data": all_passwords,
            "total_accounts_generated": len(all_passwords)
        }).execute()
    except Exception as e:
        logger.warning(f"Failed to save passwords: {e}")

    return {"accounts": accounts, "total": len(accounts), "students": len(students), "teachers": len(teachers)}


# ============================================
# الإعدادات
# ============================================
@router.get("/settings")
async def get_settings(current_user: dict = Depends(_require_admin), db=Depends(get_db)):
    result = db.table("user_settings").select("*").eq("user_id", current_user["id"]).execute()
    settings = result.data[0] if result.data else {}
    return {
        "theme": settings.get("theme", "dark"),
        "language": settings.get("language", "ar"),
        "notifications_enabled": settings.get("notifications_enabled", True),
        "avatar_url": current_user.get("avatar_url", ""),
        "email": current_user.get("email", ""),
        "full_name": current_user.get("full_name", ""),
    }


@router.put("/settings")
async def update_settings(body: dict, current_user: dict = Depends(_require_admin), db=Depends(get_db)):
    valid_langs = {"ar", "en", "de", "fr", "zh", "es"}
    valid_themes = {"dark", "light", "library", "neon"}

    theme = body.get("theme", "dark")
    if theme not in valid_themes: theme = "dark"
    lang = body.get("language", "ar")
    if lang not in valid_langs: lang = "ar"
    data = {
        "user_id": current_user["id"],
        "theme": theme,
        "language": lang,
        "notifications_enabled": bool(body.get("notifications_enabled", True)),
    }
    try:
        existing = db.table("user_settings").select("id").eq("user_id", current_user["id"]).execute()
        if existing.data:
            db.table("user_settings").update(data).eq("user_id", current_user["id"]).execute()
        else:
            db.table("user_settings").insert(data).execute()
    except Exception as e:
        logger.warning(f"Admin settings save failed: {e}")
        try:
            db.table("user_settings").upsert({"user_id": current_user["id"], "theme": theme}).execute()
        except Exception:
            pass

    if body.get("avatar_url") is not None:
        try:
            db.table("users").update({"avatar_url": body["avatar_url"]}).eq("id", current_user["id"]).execute()
        except Exception as e:
            logger.warning(f"avatar update failed: {e}")

    return {"message": "✅ تم حفظ الإعدادات"}


# ============================================================
# 🏫 ميزات الإداري المتقدمة
# ============================================================

@router.get("/school-pulse")
async def school_pulse(current_user: dict = Depends(_require_admin), db=Depends(get_db)):
    """نبض المدرسة لحظياً"""
    sid = current_user.get("school_id")
    if not sid:
        return {
            "students": 0, "teachers": 0, "active_today": 0, "active_week": 0,
            "ai_calls_week": 0, "engagement_pct": 0, "error": "no_school"
        }

    from datetime import datetime, timedelta, timezone
    today = datetime.now(timezone.utc).date().isoformat()
    week = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()

    students_count = teachers_count = active_today = active_week = ai_calls = 0
    total = 0

    try:
        users = db.table("users").select("id, role").eq("school_id", sid).execute()
        all_u = users.data or []
        total = len(all_u)
        students_count = sum(1 for u in all_u if u.get("role") == "student")
        teachers_count = sum(1 for u in all_u if u.get("role") == "teacher")
    except Exception as e:
        logger.error(f"School pulse users failed: {e}")

    # نشاط من last_login (لو العمود موجود)
    try:
        users_ll = db.table("users").select("id, last_login").eq("school_id", sid).execute()
        for u in (users_ll.data or []):
            ll = u.get("last_login")
            if ll:
                ll_s = str(ll)
                if ll_s >= today: active_today += 1
                if ll_s >= week: active_week += 1
    except Exception:
        pass

    # Fallback من analytics
    if active_week == 0:
        try:
            logs = db.table("analytics").select("student_id, created_at").eq("school_id", sid).eq("event_type", "login").gte("created_at", week).execute()
            seen_today, seen_week = set(), set()
            for l in (logs.data or []):
                sid_ = l.get("student_id")
                if not sid_: continue
                seen_week.add(sid_)
                if str(l.get("created_at", "")) >= today:
                    seen_today.add(sid_)
            active_today = max(active_today, len(seen_today))
            active_week = max(active_week, len(seen_week))
        except Exception:
            pass

    try:
        a = db.table("analytics").select("id", count="exact").eq("school_id", sid).gte("created_at", week).execute()
        ai_calls = a.count or 0
    except Exception:
        pass

    return {
        "students": students_count,
        "teachers": teachers_count,
        "active_today": active_today,
        "active_week": active_week,
        "ai_calls_week": ai_calls,
        "engagement_pct": round((active_week / max(total, 1)) * 100, 1),
    }


@router.post("/announcement")
async def make_announcement(
    body: dict,
    current_user: dict = Depends(_require_admin),
    db=Depends(get_db),
):
    """إعلان مدرسي"""
    title = (body.get("title") or "").strip()[:200]   # حد العنوان 200 حرف
    content = (body.get("content") or "").strip()[:2000]  # حد المحتوى 2000 حرف
    if not title or not content:
        raise HTTPException(400, "أدخل العنوان والمحتوى")
    try:
        db.table("analytics").insert({
            "student_id": current_user["id"],
            "school_id": current_user.get("school_id"),
            "event_type": "school_announcement",
            "event_data": {"title": title, "content": content},
        }).execute()
    except Exception:
        pass
    return {"message": "✅ تم نشر الإعلان"}


@router.get("/announcements")
async def list_announcements(current_user: dict = Depends(_require_admin), db=Depends(get_db)):
    try:
        sid = current_user.get("school_id")
        res = db.table("analytics").select("event_data, created_at").eq(
            "school_id", sid
        ).eq("event_type", "school_announcement").order("created_at", desc=True).limit(20).execute()
        return res.data or []
    except Exception:
        return []


@router.post("/incident-report")
async def incident_report(
    body: dict,
    current_user: dict = Depends(_require_admin),
):
    """مساعد كتابة تقارير حوادث/سلوك بـ AI"""
    summary = (body.get("summary") or "").strip()[:3000]
    incident_type = str(body.get("type", "سلوكي"))[:50]
    if not summary:
        raise HTTPException(400, "اكتب ملخص الحادثة")
    from app.services.ai_service import chat_with_gemini
    prompt = f"""اكتب تقرير {incident_type} مدرسي رسمي ومنظم مبني على المعلومات التالية:
{summary}

التقرير يجب أن يحتوي:
- التاريخ والوقت (ضع [اكتب التاريخ])
- الأطراف المعنية
- وصف الحادثة بشكل موضوعي
- الإجراءات المتخذة
- التوصيات
بنبرة رسمية محايدة."""
    try:
        text, _ = await chat_with_gemini(prompt, None, full_name=current_user.get("full_name", ""), role="admin")
        return {"report": text}
    except Exception:
        return {"report": "تعذر التوليد"}
